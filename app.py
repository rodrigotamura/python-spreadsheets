# Let's create a new column that is the 90% off from current value
# Let's add a chart to analise the values progression

import openpyxl as xl #This is a popular library that manage Excel Spreadsheets
from openpyxl.chart import BarChart, Reference

#openning spreadsheet file
wb = xl.load_workbook("transactions.xlsx")

#getting the sheet we'll work with
sheet = wb['Sheet1']

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = float(cell.value) * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

#creating chart
values = Reference(sheet, min_row = 2, max_row = sheet.max_row,
                   min_col=4, max_col=4)

chart = BarChart() # Creating chart object
chart.add_data(values) # Setting values on chart
sheet.add_chart(chart, 'E2') # insert chart on E2 cell
wb.save('transactions2.xlsx') # Saving file modification with another name